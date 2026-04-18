"""
PMI Dispatcher — Core Engine v3
- Rotation read from uploaded file (no hardcoding)
- Schedule-aware load balancing
- Failsafe flagging for unknown/zero-hour items
"""

import re
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# ── Constants ─────────────────────────────────────────────────────────────────
MECHANIC_ORDER = [
    "Frank", "Santos", "Thomas", "Ben", "Hugo",
    "Robert", "Devin", "Steven", "Rafa", "Brian", "Cesar"
]

DAYS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

# Sort schedule bundled with app — matches Sort Equp. usage data
# Tuples are (start_hour_24h, end_hour_24h); end > 24 means crosses midnight
SORT_SCHEDULE = {
    "Preload":  {"Sun": None,      "Mon": (3, 7),   "Tue": (4, 9),   "Wed": (4, 9),
                 "Thu": (4, 9),    "Fri": (4, 9),   "Sat": (4, 9)},
    "Day":      {"Sun": (18, 24),  "Mon": (13, 16), "Tue": (13, 16), "Wed": (13, 16),
                 "Thu": (13, 16),  "Fri": None,     "Sat": (17, 24)},
    "Twilight": {"Sun": None,      "Mon": (16, 22), "Tue": (16, 22), "Wed": (16, 22),
                 "Thu": (16, 22),  "Fri": (16, 22), "Sat": None},
    "Night":    {"Sun": None,      "Mon": (23, 27), "Tue": (23, 27), "Wed": (23, 27),
                 "Thu": (23, 27),  "Fri": (23, 27), "Sat": None},
}

# Downtime score per sort overlap combination
# Higher score = more downtime available for heavy PMIs
DOWNTIME_SCORES = {
    frozenset(["Preload"]):                       5.0,
    frozenset(["Preload", "Day"]):                4.0,
    frozenset(["Day"]):                           3.5,
    frozenset(["Day", "Twilight"]):               3.0,
    frozenset(["Preload", "Day", "Twilight"]):    2.5,
    frozenset(["Day", "Twilight", "Night"]):      1.5,
    frozenset(["Twilight"]):                      2.0,
    frozenset(["Twilight", "Night"]):             1.0,
    frozenset(["Night"]):                         2.0,
}

PM_TYPE_ORDER = ["PMI-04", "UNKNOWN", "PMI-05", "PMI-06", "PMI-06A", "PMI-02", "PMI-10"]

# ── Quarter detection ─────────────────────────────────────────────────────────
def quarter_from_month(month: int) -> int:
    return (month - 1) // 3 + 1


# ── Rotation file parser ──────────────────────────────────────────────────────
def parse_rotation(file_obj) -> dict:
    """
    Read rotation from uploaded xlsx.
    Required columns: Mechanic, Sorter, Walk1, [Walk2], Tractor, [Special1..N]
    Returns: {mechanic: {sorter, walks, tractor, special}}
    """
    df = pd.read_excel(file_obj, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=[df.columns[0]])

    rotation = {}
    for _, row in df.iterrows():
        mech = str(row.iloc[0]).strip()
        if not mech or mech.lower() in ['mechanic', 'name', 'nan']:
            continue
        # Only process rows where name matches a known mechanic
        if mech not in MECHANIC_ORDER:
            continue

        sorter = _clean(row.get('Sorter', ''))

        walks = []
        for col in df.columns:
            if re.match(r'Walk\d*', col, re.IGNORECASE):
                v = _clean(row.get(col, ''))
                if v:
                    walks.append(v)

        tractor_raw = _clean(row.get('Tractor', ''))
        tractors = [p.strip() for p in re.split(r'[/,]', tractor_raw) if p.strip()] if tractor_raw else []

        specials = []
        for col in df.columns:
            if re.match(r'Special\d*', col, re.IGNORECASE):
                v = _clean(row.get(col, '')).lower()
                if v:
                    specials.append(v)

        rotation[mech] = {
            'sorter': sorter,
            'walks': walks,
            'tractor': tractors,
            'special': specials,
        }

    return rotation


def _clean(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s


# ── Schedule file parser ──────────────────────────────────────────────────────
def parse_schedule(file_obj) -> dict:
    """
    Read mechanic schedule from uploaded xlsx.
    Required columns: Mechanic, Sun-Sat (shift times), SortOverlap1/2/3
    Returns: {mechanic: {shifts, sort_overlaps, downtime_score}}
    """
    df = pd.read_excel(file_obj, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=[df.columns[0]])

    schedule = {}
    for _, row in df.iterrows():
        mech = str(row.iloc[0]).strip()
        if not mech or mech.lower() in ['mechanic', 'name', 'nan']:
            continue
        if mech not in MECHANIC_ORDER:
            continue

        shifts = {}
        for day in DAYS:
            val = _clean(row.get(day, ''))
            shifts[day] = val if val and val.lower() != 'off' else 'Off'

        overlaps = []
        for col in df.columns:
            if re.match(r'SortOverlap\d*', col, re.IGNORECASE) or col == 'Sort Overlap':
                v = _clean(row.get(col, ''))
                if v:
                    overlaps.append(v.strip().rstrip())

        overlap_key = frozenset(overlaps)
        downtime_score = DOWNTIME_SCORES.get(overlap_key, max(1.0, 4.0 - len(overlaps)))

        schedule[mech] = {
            'shifts': shifts,
            'sort_overlaps': overlaps,
            'downtime_score': downtime_score,
        }

    return schedule


# ── PMI type extraction ───────────────────────────────────────────────────────
def extract_pmi_type(desc: str) -> str:
    desc_upper = str(desc).upper()
    if "PDC" in desc_upper and "ELECTRICAL" in desc_upper:
        return "PMI-04"
    m = re.search(r'PMI-\d+[A-Z]?', str(desc), re.IGNORECASE)
    return m.group(0).upper() if m else "UNKNOWN"


# ── Sorter matcher ────────────────────────────────────────────────────────────
def _sorter_pat(label: str) -> re.Pattern:
    m = re.match(r'^([A-Za-z]+)(.*)', label)
    if not m:
        return re.compile(re.escape(label), re.IGNORECASE)
    prefix, rest = m.group(1), m.group(2)
    rest_pat = re.sub(r'[/\-\s]', r'[/\\- ]', re.escape(rest))
    return re.compile(r'(?<![A-Za-z0-9])' + re.escape(prefix) + rest_pat, re.IGNORECASE)


def match_sorter(desc: str, rotation: dict) -> str | None:
    for mech, cfg in rotation.items():
        if cfg['sorter'] and _sorter_pat(cfg['sorter']).search(desc):
            return mech
        for sp in cfg.get('special', []):
            if sp == 'ps3' and re.search(r'(?<![A-Za-z0-9])PS3', desc, re.IGNORECASE):
                return mech
    return None


# ── Walk matcher ──────────────────────────────────────────────────────────────
def match_walk(desc: str, rotation: dict) -> str | None:
    for mech, cfg in rotation.items():
        for w in cfg.get('walks', []):
            if re.search(r'(?<!\d)' + re.escape(w) + r'(?!\d)', desc, re.IGNORECASE):
                return mech
    return None


# ── Tractor matcher ───────────────────────────────────────────────────────────
def build_tractor_patterns(rotation: dict) -> dict:
    patterns = {}
    for mech, cfg in rotation.items():
        pats = []
        for label in cfg.get('tractor', []):
            if re.match(r'PE', label, re.IGNORECASE):
                num = re.sub(r'[^0-9]', '', label)
                pats.append(re.compile(r'PE[\s\-]?' + num + r'\b', re.IGNORECASE))
                pats.append(re.compile(r'IRREG[\s\-]?TRACTOR.*PE[\s\-]?' + num, re.IGNORECASE))
            elif label.isdigit():
                pats.append(re.compile(r'IRREG[\s\-]?TRACTOR[\s\-#]*' + label + r'(?!\d)', re.IGNORECASE))
                pats.append(re.compile(r'\bTUG[\s\-#]*' + label + r'(?!\d)', re.IGNORECASE))
        patterns[mech] = pats
    return patterns


def match_tractor(desc: str, tractor_patterns: dict) -> str | None:
    if not re.search(r'IRREG[\s\-]?TRACTOR|TUG\s*\d|PE[\s\-]?\d', desc, re.IGNORECASE):
        return None
    for mech, pats in tractor_patterns.items():
        for p in pats:
            if p.search(desc):
                return mech
    return None


# ── Special token matcher ─────────────────────────────────────────────────────
def match_special(desc: str, rotation: dict, pmi_type: str) -> str | None:
    if pmi_type not in ("PMI-04", "PMI-05", "UNKNOWN"):
        return None

    desc_lower = desc.lower()

    if re.search(r'power\s+turn', desc_lower):
        target = "old pt walk" if re.search(r'\bold\b', desc_lower) else \
                 ("new pt walk" if re.search(r'\bnew\b', desc_lower) else None)
        if target:
            for mech, cfg in rotation.items():
                if any(target in sp for sp in cfg.get('special', [])):
                    return mech

    if re.search(r'pdc|electrical.{0,10}connection', desc_lower):
        for mech, cfg in rotation.items():
            if any("pdc" in sp or "electrical" in sp for sp in cfg.get('special', [])):
                return mech

    if re.search(r'eyewash|emergency\s+shower', desc_lower):
        for mech, cfg in rotation.items():
            if any("eyewash" in sp or "emergency" in sp for sp in cfg.get('special', [])):
                return mech

    if pmi_type == "PMI-04" and re.search(r'\bsslaw\b|w-sslaw', desc_lower):
        for mech, cfg in rotation.items():
            if any("sslaw" in sp for sp in cfg.get('special', [])):
                return mech

    if pmi_type == "PMI-05" and re.search(r'\bsstt', desc_lower):
        for mech, cfg in rotation.items():
            if any("sstt" in sp for sp in cfg.get('special', [])):
                return mech

    if pmi_type in ("PMI-04", "PMI-05") and re.search(r'\bm7', desc_lower):
        for mech, cfg in rotation.items():
            if any(sp == "m7" for sp in cfg.get('special', [])):
                return mech

    if pmi_type == "PMI-04" and re.search(r'm.?exit\s+light|exit\s+light', desc_lower):
        for mech, cfg in rotation.items():
            if any("exit" in sp for sp in cfg.get('special', [])):
                return mech

    return None


# ── Assignment engine ─────────────────────────────────────────────────────────
def assign_mechanic(row: pd.Series, rotation: dict, tractor_patterns: dict) -> tuple:
    """
    Returns (mechanic, flag_reason) where flag_reason is None if clean.
    PMI-06/10/02 always go to balancer (return Unassigned).
    """
    desc = str(row["Description"])
    pmi_type = row["PMI_Type"]

    # All non-04/05 → load balancer
    if pmi_type in ("PMI-06", "PMI-06A", "PMI-10", "PMI-02"):
        return "Unassigned", None

    try:
        m = match_special(desc, rotation, pmi_type)
        if m:
            return m, None

        if pmi_type == "UNKNOWN":
            result = match_walk(desc, rotation) or match_sorter(desc, rotation)
            if result:
                return result, None
            return "Unassigned", "PMI type unknown and no pattern match — verify description"

        if pmi_type == "PMI-04":
            result = match_walk(desc, rotation) or match_sorter(desc, rotation)
            if result:
                return result, None
            return "Unassigned", "PMI-04 with no walk or sorter match — verify description"

        if pmi_type == "PMI-05":
            result = match_sorter(desc, rotation) or match_tractor(desc, tractor_patterns)
            if result:
                return result, None
            return "Unassigned", "PMI-05 with no sorter or tractor match — verify description"

    except Exception as e:
        return "Unassigned", f"Error during assignment: {e}"

    return "Unassigned", "No rule matched"


# ── Master plan loader ────────────────────────────────────────────────────────
def _parse_hours(h) -> float | None:
    m = re.match(r'(\d+):(\d+)', str(h))
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60
    try:
        return float(h)
    except:
        return None


def _parse_num_pms(val) -> int | None:
    if val is None:
        return None
    if hasattr(val, 'day'):
        return val.day
    if isinstance(val, (int, float)) and not pd.isna(val):
        return max(1, int(val))
    m = re.search(r'\d+', str(val))
    return int(m.group(0)) if m else None


def _norm(desc: str) -> str:
    return re.sub(r'\s+', ' ', str(desc).strip().lower())


def load_master_plan(file_path: str, month_index: int = 0) -> dict:
    df = pd.read_excel(file_path, sheet_name=month_index, header=None)
    header_row = None
    for i, row in df.iterrows():
        if any("Description" in str(v) for v in row.values):
            header_row = i
            break
    if header_row is None:
        return {}

    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["Description"])

    lookup = {}
    for _, row in df.iterrows():
        desc = _clean(row.get("Description", ""))
        if not desc:
            continue
        pmi_type = str(row.get("Type", "")).strip().upper()
        hours_dec = _parse_hours(row.get("Estimated\nHours"))
        if hours_dec is None:
            continue
        if "PMI-04" in pmi_type:
            num_pms = _parse_num_pms(row.get("# of\nPMs"))
            if num_pms and num_pms > 1:
                hours_dec = hours_dec / num_pms
        key = _norm(desc)
        if key not in lookup:
            lookup[key] = hours_dec

    return lookup


# ── Compliance file loader ────────────────────────────────────────────────────
def load_maximo_export(file_obj) -> pd.DataFrame:
    raw = pd.read_excel(file_obj, sheet_name=0, header=None)
    header_row = None
    for i, row in raw.iterrows():
        if any("WO Number" in str(v) for v in row.values):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Cannot find 'WO Number' header in compliance file.")

    df = pd.read_excel(file_obj, sheet_name=0, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[["WO Number", "Description", "PM Due Date"]].copy()
    df = df.dropna(subset=["WO Number", "Description"])
    df["Description"] = df["Description"].fillna("").astype(str).str.strip()
    df["WO Number"] = df["WO Number"].astype(str).str.strip()
    df["PM Due Date"] = pd.to_datetime(df["PM Due Date"], errors="coerce")
    df = df.dropna(subset=["PM Due Date"])
    return df


def filter_by_month(df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    mask = (df["PM Due Date"].dt.year == year) & (df["PM Due Date"].dt.month == month)
    return df[mask].copy()


# ── Schedule-aware greedy balancer ────────────────────────────────────────────
def balance_unowned(unassigned: pd.DataFrame, assigned_hours: dict,
                    master_hours: dict, schedule: dict | None = None) -> tuple:
    """
    Returns (balanced_df, final_hours, flags_list)
    flags_list: items that need manual review
    """
    result = unassigned.copy()
    result["Mechanic"] = "Unassigned"
    hours = dict(assigned_hours)
    flags = []

    mechs = MECHANIC_ORDER

    # Build downtime scores
    if schedule:
        downtime = {m: schedule.get(m, {}).get('downtime_score', 2.0) for m in mechs}
    else:
        downtime = {m: 2.0 for m in mechs}

    def get_hrs(row):
        key = _norm(row["Description"])
        h = master_hours.get(key, 0.0)
        return h

    result["_h"] = result.apply(get_hrs, axis=1)
    result = result.sort_values("_h", ascending=False).reset_index(drop=True)

    for idx, row in result.iterrows():
        desc = str(row["Description"])
        pmi_type = row["PMI_Type"]
        h = row["_h"]

        # Flag zero-hour items (still assign, but note it)
        if h == 0:
            flags.append({
                'WO Number': row["WO Number"],
                'Description': desc,
                'PMI Type': pmi_type,
                'Flag': 'Hours not found in master plan — verify labor estimate',
                'Assigned To': '(pending)'
            })

        is_qt = bool(re.match(r'QT-\d+', desc.strip(), re.IGNORECASE))

        if is_qt:
            # QT = sort-compatible, pure hours balance
            best = min(mechs, key=lambda m: hours.get(m, 0))
        else:
            # Downtime-required: blend hours + downtime penalty
            # Heavy PMs get stronger downtime weighting
            weight = 2.5 if h > 2.0 else 0.75

            def score(m):
                return hours.get(m, 0) + (4.0 - downtime.get(m, 2.0)) * weight

            best = min(mechs, key=score)

        result.at[idx, "Mechanic"] = best
        hours[best] = hours.get(best, 0) + h

        # Update flag with assigned mechanic
        if h == 0 and flags and flags[-1]['WO Number'] == row["WO Number"]:
            flags[-1]['Assigned To'] = best

    result = result.drop(columns=["_h"])
    return result, hours, flags


# ── Full dispatch pipeline ────────────────────────────────────────────────────
def dispatch(df: pd.DataFrame, master_hours: dict, rotation: dict,
             schedule: dict | None = None) -> tuple:
    """
    Returns: (blocks, hour_summary, all_flags)
    blocks: {mechanic: DataFrame}
    """
    tractor_patterns = build_tractor_patterns(rotation)

    df = df.copy()
    df["PMI_Type"] = df["Description"].apply(extract_pmi_type)

    assignment_results = df.apply(
        lambda r: assign_mechanic(r, rotation, tractor_patterns), axis=1
    )
    df["Mechanic"] = assignment_results.apply(lambda x: x[0])
    df["_flag"] = assignment_results.apply(lambda x: x[1])

    # Collect assignment flags (unmatchable PMI-04/05)
    assign_flags = []
    flagged = df[df["_flag"].notna() & (df["Mechanic"] == "Unassigned")]
    for _, row in flagged.iterrows():
        assign_flags.append({
            'WO Number': row["WO Number"],
            'Description': row["Description"],
            'PMI Type': row["PMI_Type"],
            'Flag': row["_flag"],
            'Assigned To': 'Unassigned — review required'
        })

    assigned = df[df["Mechanic"] != "Unassigned"].copy()
    unassigned = df[df["Mechanic"] == "Unassigned"].copy()

    # Tally hours from owned assignments
    assigned_hours = {}
    for _, row in assigned.iterrows():
        h = master_hours.get(_norm(row["Description"]), 0)
        mech = row["Mechanic"]
        assigned_hours[mech] = assigned_hours.get(mech, 0) + h

    # Balance unowned
    balance_flags = []
    if len(unassigned) > 0:
        balanced, final_hours, balance_flags = balance_unowned(
            unassigned, assigned_hours, master_hours, schedule
        )
        all_df = pd.concat([assigned, balanced], ignore_index=True)
    else:
        all_df = assigned
        final_hours = assigned_hours

    all_flags = assign_flags + balance_flags

    # Format output
    all_df["PM Due Date Fmt"] = all_df["PM Due Date"].dt.strftime("%-m/%-d/%y")
    out_cols = ["WO Number", "Description", "PM Due Date Fmt", "PMI_Type"]
    rename = {"PM Due Date Fmt": "PM Due Date"}

    blocks = {}
    for mech in MECHANIC_ORDER + ["Unassigned"]:
        subset = all_df[all_df["Mechanic"] == mech][out_cols].copy().rename(columns=rename)
        subset["_st"] = subset["PMI_Type"].apply(
            lambda x: PM_TYPE_ORDER.index(x) if x in PM_TYPE_ORDER else len(PM_TYPE_ORDER))
        subset["_sd"] = all_df.loc[subset.index, "PM Due Date"]
        subset = subset.sort_values(["_st", "_sd", "WO Number"]).drop(columns=["_st", "_sd"])
        if not subset.empty:
            blocks[mech] = subset.reset_index(drop=True)

    return blocks, final_hours, all_flags


# ── Excel writer ──────────────────────────────────────────────────────────────
def write_excel(blocks: dict, hour_summary: dict, flags: list, output_path: str) -> None:
    wb = Workbook()

    # ── Sheet 1: Final Master ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Final Master"

    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    col_widths = [14, 55, 14, 12]
    headers = ["WO Number", "Description", "PM Due Date", "PMI Type"]

    first = True
    for mech, df in blocks.items():
        if mech == "Unassigned":
            continue
        if not first:
            ws.append([])
        first = False
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=mech)
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = green_fill
            ws.cell(row=r, column=c).font = white_bold
        hr = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            ws.cell(row=hr, column=c, value=h).font = bold_font
        for _, row in df.iterrows():
            ws.append([row["WO Number"], row["Description"],
                       row["PM Due Date"], row["PMI_Type"]])

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Hour Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Hour Summary")
    ws2.append(["Mechanic", "Assigned Hours", "PM Count"])
    for c in range(1, 4):
        ws2.cell(1, c).font = bold_font
    for mech in MECHANIC_ORDER:
        ws2.append([mech, round(hour_summary.get(mech, 0), 2),
                    len(blocks.get(mech, pd.DataFrame()))])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: Review Required (flags) ──────────────────────────────────────
    if flags:
        ws3 = wb.create_sheet("Review Required")
        flag_headers = ["WO Number", "Description", "PMI Type", "Flag", "Assigned To"]
        ws3.append(flag_headers)
        for c in range(1, len(flag_headers) + 1):
            ws3.cell(1, c).font = bold_font
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for flag in flags:
            row_data = [flag.get(h, '') for h in flag_headers]
            ws3.append(row_data)
            r = ws3.max_row
            for c in range(1, len(flag_headers) + 1):
                ws3.cell(r, c).fill = yellow_fill
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 50
        ws3.column_dimensions["C"].width = 12
        ws3.column_dimensions["D"].width = 55
        ws3.column_dimensions["E"].width = 18

    wb.save(output_path)


# ── Validation helpers ────────────────────────────────────────────────────────
def validate_rotation(rotation: dict) -> list:
    """Returns list of error strings, empty if valid."""
    errors = []
    if not rotation:
        return ["Rotation file is empty or could not be parsed."]
    for mech in MECHANIC_ORDER:
        if mech not in rotation:
            errors.append(f"Mechanic '{mech}' not found in rotation file.")
        else:
            cfg = rotation[mech]
            if not cfg.get('sorter'):
                errors.append(f"{mech}: missing Sorter assignment.")
            if not cfg.get('walks'):
                errors.append(f"{mech}: missing Walk assignment(s).")
    return errors


def validate_schedule(schedule: dict) -> list:
    errors = []
    if not schedule:
        return ["Schedule file is empty or could not be parsed."]
    for mech in MECHANIC_ORDER:
        if mech not in schedule:
            errors.append(f"Mechanic '{mech}' not found in schedule file.")
    return errors


# ── Main runner ───────────────────────────────────────────────────────────────
def run(compliance_path: str, rotation_obj, schedule_obj,
        master_path: str, year: int, month: int, output_path: str) -> dict:

    quarter = quarter_from_month(month)
    master_sheet_index = month - 1

    with open(compliance_path, "rb") as f:
        df = load_maximo_export(f)
    df = filter_by_month(df, year, month)

    rotation = parse_rotation(rotation_obj)
    schedule = parse_schedule(schedule_obj)
    master_hours = load_master_plan(master_path, master_sheet_index)

    blocks, hour_summary, flags = dispatch(df, master_hours, rotation, schedule)

    assigned = sum(len(v) for k, v in blocks.items() if k != "Unassigned")
    unassigned = len(blocks.get("Unassigned", pd.DataFrame()))

    write_excel(blocks, hour_summary, flags, output_path)

    return {
        'total': len(df),
        'assigned': assigned,
        'unassigned': unassigned,
        'flags': len(flags),
        'quarter': quarter,
        'hour_summary': {m: round(hour_summary.get(m, 0), 2) for m in MECHANIC_ORDER},
    }
